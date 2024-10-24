from flask import request, redirect, flash, abort, url_for
from flask_login import login_required
from werkzeug.utils import secure_filename

import os
import openpyxl
from datetime import datetime
import pytz

from app import app
from app import db

, only_admin
from app.Forms.Global import IMPORTForm
from app.misc.get_model import models
from app.misc.gen_id import generate_id

@app.route("/importacao_usuário", methods=["POST"])
@login_required


def importacao_usuário():

    try:
        form = IMPORTForm()

        if form.validate_on_submit():

            importe_epi(form)
            return redirect(f'{request.referrer}')

        if form.errors:

            for erros in list(form.errors):
                message = form.errors[erros][0]
                flash(message, "error")

    except Exception as e:
        abort(500)


def importe_epi(form):

    doc =  form.arquivo.data
    docname = secure_filename(doc.filename)

    doc.save(os.path.join(os.getcwd(), "app", 'Temp', f"{docname}"))
    doc_path = os.path.join(os.getcwd(), "app", 'Temp', f"{docname}")
    
    update_counts(doc_path)
    

def update_counts(doc_path: str):
    
    model = models("users")
    wb = openpyxl.load_workbook(filename=doc_path)
    ws = wb.active

    for i in range(2, ws.max_row+1):
        
        kwargs = {}
        
        nomes_colunas = ["id", "login", "nome_usuario", "type_user", "email", "password", "license_key"]
        for index in range(0, len(nomes_colunas)):
            for nome_coluna in nomes_colunas:
                column_value = index+1
                nome_coluna_planilha = ws.cell(row=1, column=column_value).value
                if nome_coluna_planilha is None:
                    continue
                elif nome_coluna.upper() == str(nome_coluna_planilha).upper():
                    valor_celula = ws.cell(row=i, column=column_value).value
                    if valor_celula is None:
                        kwargs[nome_coluna] = ''
                    else:
                        kwargs[nome_coluna] = valor_celula
                        break
        kwargs["login_id"] = generate_id()
        record = model.query.filter_by(id = kwargs.get("login")).first()
        if record is None:    
            new_record = model(**kwargs)
            db.session.add(new_record)
            db.session.commit()

        if i == ws.max_row:
            
            flash("Usuários importados com sucesso!", "success")
            return redirect(url_for("config"))

