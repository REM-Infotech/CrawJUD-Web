import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from pytz import timezone

from app.misc.MakeTemplate.appends import listas


class MakeXlsx:

    def __init__(self, model_name: str, display_name: str):
        """

        ### type: Tipo da planilha (sucesso, erro)
        ### system: o sistema que está sendo executado a automação Ex.: PROJUDI, ESAJ, ELAW, ETC.

        """
        self.model_name = model_name
        self.displayname = display_name
        self.listas = listas()

    def make_output(self):

        temp_dir = os.path.join(os.getcwd(), "Temp")
        os.makedirs(temp_dir, exist_ok=True)
        name_file = f"{self.displayname.upper()} - {datetime.now(timezone('Etc/GMT+4')).strftime('%H-%M-%S')}.xlsx"

        path_template = os.path.join(temp_dir, name_file)

        # Criar um novo workbook e uma planilha
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Resultados", 0)
        sheet = workbook.active

        # Cabeçalhos iniciais
        cabecalhos = ["NUMERO_PROCESSO"]
        list_to_append = []

        itens_append = self.listas(self.model_name)
        if itens_append:
            list_to_append.extend(itens_append)

        elif not itens_append:
            itens_append = self.listas(self.model_name.split("_")[0])
            if itens_append:
                list_to_append.extend(itens_append)

        cabecalhos.extend(list_to_append)
        # Definir estilo
        my_red = openpyxl.styles.colors.Color(rgb="A6A6A6")
        my_fill = PatternFill(patternType="solid", fgColor=my_red)
        bold_font = Font(name="Times New Roman", italic=True)

        # Escrever os cabeçalhos na primeira linha da planilha e aplicar estilo
        for pos, coluna in enumerate(cabecalhos):
            item = sheet.cell(row=1, column=pos + 1, value=coluna.upper())
            item.font = bold_font
            item.fill = my_fill

        # Issue: [B110:try_except_pass] Try, Except, Pass detected.
        # Severity: Low   Confidence: High
        # CWE: CWE-703 (https://cwe.mitre.org/data/definitions/703.html)
        # vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv

        # for col in sheet.columns:
        #     max_length = 0
        #     column = col[0].column_letter  # Get the column name
        #     for cell in col:
        #         try:  # Necessary to avoid error on empty cells
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(str(cell.value))
        #         except Exception:
        #             pass
        #     adjusted_width = (max_length + 2) * 1.2
        #     sheet.column_dimensions[column].width = adjusted_width

        # workbook.save(path_template)
        # return path_template, name_file

        """Ajustar a largura das colunas"""
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:

                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(path_template)
        return path_template, name_file
